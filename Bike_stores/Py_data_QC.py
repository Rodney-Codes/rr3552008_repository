#importing necessary libraries
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#adding relevant workbook to code
wb = load_workbook('C:\\Code\\SQL projects\\Bike store database\\temp_py_code.xlsx')
wb.active = wb['Sheet1']
ws = wb.active

#checking if the data file has empty cells or not
f = 0
for row in range(2,ws.max_row+1):
    for col in range(1,ws.max_column+1):
        char = get_column_letter(col)
        if ws[char + str(row)].value is None:
            print ("Data file has empty cells")
            f=1
            break
    if f == 1:
        break

#checking if key data is empty or not
f = 0
temp_list = ["product_id", "brand_id", "category_id", "product_list_price", "order_id", "order_item_quantity", "order_item_list_price", "discount", "customer_id", "store_id", "staff_id", "customer_first_name", "customer_last_name"]
for col in range(1,ws.max_column+1):
    for row in range(2,ws.max_row+1):
        char = get_column_letter(col)
        if ws[char + str(1)].value in temp_list and ws[char + str(row)].value is None:
            print ("Data file is missing crucial information")
            f=1
            break
    if f == 1:
        break
if f == 0:
    print ("Data file has all the necessary information")

#logical checks
for col in range(1,ws.max_column+1):
    if ws.cell(row=1, column=col).value == "order_item_quantity":
        order_quantity = get_column_letter(col)
    elif ws.cell(row=1, column=col).value == "store_quantity":
        store_quantity = get_column_letter(col)
    elif ws.cell(row=1, column=col).value == "order_status":
        order_status = get_column_letter(col)
    elif ws.cell(row=1, column=col).value == "order_id":
        order_id = get_column_letter(col)
    elif ws.cell(row=1, column=col).value == "product_name":
        product_name = get_column_letter(col)
for row in range(2,ws.max_row+1):
    if ws[order_quantity + str(1)].value > ws[store_quantity + str(1)].value or ws[order_status + str(row)].value != 4:
        print ("Order cannot be completed for" + str(ws[order_id + str(row)].value) + " ordering " + str(ws[product_name + str(row)].value))

#range checks
for col in range(1,ws.max_column+1):
    if ws.cell(row=1, column=col).value == "discount":
        discount = get_column_letter(col)
    elif ws.cell(row=1, column=col).value == "product_id":
        product_id = get_column_letter(col)
for row in range(2,ws.max_row+1):
    if float(ws[discount + str(row)].value) > 1:
        print ("Incorrect discount for the product " + str(ws[product_id + str(row)].value) + " is given")

#text or number checks
f = 0
num_list = ["product_id", "brand_id", "category_id", "model_year", "product_list_price", "order_id", "item_id", "order_item_quantity", "order_item_list_price", "discount", "customer_id", "order_status", "store_id", "staff_id", "customer_zip_code", "store_zip_code", "active", "manager_id", "store_quantity"]
for col in range(1,ws.max_column+1):
    for row in range(2,ws.max_row+1):
        char = get_column_letter(col)
        if ws[char + str(row)].value is not None:
            if ws[char + str(1)].value in num_list and isinstance(ws[char + str(row)].value,(int, float)) == False:
                print ("Data file has incorrect data types")
                f = 1
                break
            elif ws[char + str(1)].value in num_list and isinstance(ws[char + str(row)].value,str) == False:
                print ("Data file has incorrect data types")
                f = 1
                break
    if f == 1:
        break
if f == 0:
    print ("Data file has correct datatypes")
